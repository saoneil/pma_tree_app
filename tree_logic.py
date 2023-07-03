from resources.resources import get_connection, get_dataframe
import backend
import pandas as pd
import math
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from tkinter import Tk, filedialog

def ask_save_location():
    destination_path = filedialog.asksaveasfilename(
    defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')]
    )

    if not destination_path:
        # User canceled the file dialog
        print("Save operation canceled.")
        #exit()

    return destination_path
def two_tree(competitors:list, competitors_with_byes:list):
    original_file_path = 'tree_templates\\2.xlsx'

    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook['Sheet1']

    ## entering main competitors
    sheet['D10'] = competitors[0][0]
    sheet['D11'] = competitors[0][1]

    sheet['D23'] = competitors[1][0]
    sheet['D24'] = competitors[1][1]

    ## entering the bye competitors
    sheet['J16'] = competitors_with_byes[0][0]
    sheet['J17'] = competitors_with_byes[0][1]

    save_location = ask_save_location()
    if save_location is not None:
        workbook.save(save_location)
    else:
        workbook.close()
def four_tree(competitors:list, competitors_with_byes:list):
    original_file_path = 'tree_templates\\4.xlsx'

    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook['Sheet1']

    ## entering main competitors
    sheet['C10'] = competitors[0][0]
    sheet['C11'] = competitors[0][1]

    sheet['C16'] = competitors[1][0]
    sheet['C17'] = competitors[1][1]

    sheet['C20'] = competitors[2][0]
    sheet['C21'] = competitors[2][1]

    sheet['C26'] = competitors[3][0]
    sheet['C27'] = competitors[3][1]

    ## entering the bye competitors
    sheet['G13'] = competitors_with_byes[0][0]
    sheet['G14'] = competitors_with_byes[0][1]


    save_location = ask_save_location()
    if save_location is not None:
        workbook.save(save_location)
    else:
        workbook.close()
def eight_tree(competitors:list, competitors_with_byes:list):
    original_file_path = 'tree_templates\\8.xlsx'

    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook['Sheet1']

    ## entering main competitors
    sheet['B8'] = competitors[0][0]
    sheet['B9'] = competitors[0][1]

    sheet['B36'] = competitors[1][0]
    sheet['B37'] = competitors[1][1]

    sheet['B28'] = competitors[2][0]
    sheet['B29'] = competitors[2][1]

    sheet['B16'] = competitors[3][0]
    sheet['B17'] = competitors[3][1]

    sheet['B20'] = competitors[4][0]
    sheet['B21'] = competitors[4][1]

    sheet['B24'] = competitors[5][0]
    sheet['B25'] = competitors[5][1]

    sheet['B32'] = competitors[6][0]
    sheet['B33'] = competitors[6][1]

    sheet['B12'] = competitors[7][0]
    sheet['B13'] = competitors[7][1]

    ## entering the bye competitors
    sheet['E10'] = competitors_with_byes[0][0]
    sheet['E11'] = competitors_with_byes[0][1]

    sheet['E34'] = competitors_with_byes[1][0]
    sheet['E35'] = competitors_with_byes[1][1]

    sheet['E26'] = competitors_with_byes[2][0]
    sheet['E27'] = competitors_with_byes[2][1]


    save_location = ask_save_location()
    if save_location is not None:
        workbook.save(save_location)
    else:
        workbook.close()
def sixteen_tree(competitors:list, competitors_with_byes:list):
    original_file_path = 'tree_templates\\16.xlsx'

    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook['Sheet1']

    ## entering main competitors
    sheet['B3'] = competitors[0][0]
    sheet['B4'] = competitors[0][1]

    sheet['B63'] = competitors[1][0]
    sheet['B64'] = competitors[1][1]

    sheet['B47'] = competitors[2][0]
    sheet['B48'] = competitors[2][1]

    sheet['B19'] = competitors[3][0]
    sheet['B20'] = competitors[3][1]

    sheet['B31'] = competitors[4][0]
    sheet['B32'] = competitors[4][1]

    sheet['B35'] = competitors[5][0]
    sheet['B36'] = competitors[5][1]

    sheet['B51'] = competitors[6][0]
    sheet['B52'] = competitors[6][1]

    sheet['B15'] = competitors[7][0]
    sheet['B16'] = competitors[7][1]

    sheet['B11'] = competitors[8][0]
    sheet['B12'] = competitors[8][1]

    sheet['B55'] = competitors[9][0]
    sheet['B56'] = competitors[9][1]

    sheet['B39'] = competitors[10][0]
    sheet['B40'] = competitors[10][1]

    sheet['B27'] = competitors[11][0]
    sheet['B28'] = competitors[11][1]

    sheet['B23'] = competitors[12][0]
    sheet['B24'] = competitors[12][1]

    sheet['B43'] = competitors[13][0]
    sheet['B44'] = competitors[13][1]

    sheet['B59'] = competitors[14][0]
    sheet['B60'] = competitors[14][1]

    sheet['B7'] = competitors[15][0]
    sheet['B8'] = competitors[15][1]

    ## entering the bye competitors
    sheet['E5'] = competitors_with_byes[0][0]
    sheet['E6'] = competitors_with_byes[0][1]

    sheet['E61'] = competitors_with_byes[1][0]
    sheet['E62'] = competitors_with_byes[1][1]

    sheet['E45'] = competitors_with_byes[2][0]
    sheet['E46'] = competitors_with_byes[2][1]

    sheet['E21'] = competitors_with_byes[3][0]
    sheet['E22'] = competitors_with_byes[3][1]

    sheet['E29'] = competitors_with_byes[4][0]
    sheet['E30'] = competitors_with_byes[4][1]

    sheet['E37'] = competitors_with_byes[5][0]
    sheet['E38'] = competitors_with_byes[5][1]

    sheet['E53'] = competitors_with_byes[6][0]
    sheet['E54'] = competitors_with_byes[6][1]

    save_location = ask_save_location()
    if save_location is not None:
        workbook.save(save_location)
    else:
        workbook.close()
def thirtytwo_tree(competitors:list, competitors_with_byes:list):
    original_file_path = 'tree_templates\\32.xlsx'

    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook['Sheet1']

    ## entering main competitors
    sheet['B3'] = competitors[0][0]
    sheet['B4'] = competitors[0][1]

    sheet['B127'] = competitors[1][0]
    sheet['B128'] = competitors[1][1]

    sheet['B95'] = competitors[2][0]
    sheet['B96'] = competitors[2][1]

    sheet['B35'] = competitors[3][0]
    sheet['B36'] = competitors[3][1]

    sheet['B59'] = competitors[4][0]
    sheet['B60'] = competitors[4][1]

    sheet['B71'] = competitors[5][0]
    sheet['B72'] = competitors[5][1]

    sheet['B103'] = competitors[6][0]
    sheet['B104'] = competitors[6][1]

    sheet['B27'] = competitors[7][0]
    sheet['B28'] = competitors[7][1]

    sheet['B19'] = competitors[8][0]
    sheet['B20'] = competitors[8][1]

    sheet['B111'] = competitors[9][0]
    sheet['B112'] = competitors[9][1]

    sheet['B79'] = competitors[10][0]
    sheet['B80'] = competitors[10][1]

    sheet['B51'] = competitors[11][0]
    sheet['B52'] = competitors[11][1]

    sheet['B43'] = competitors[12][0]
    sheet['B44'] = competitors[12][1]

    sheet['B87'] = competitors[13][0]
    sheet['B88'] = competitors[13][1]

    sheet['B119'] = competitors[14][0]
    sheet['B120'] = competitors[14][1]

    sheet['B11'] = competitors[15][0]
    sheet['B12'] = competitors[15][1]

    sheet['B15'] = competitors[16][0]
    sheet['B16'] = competitors[16][1]

    sheet['B115'] = competitors[17][0]
    sheet['B116'] = competitors[17][1]

    sheet['B83'] = competitors[18][0]
    sheet['B84'] = competitors[18][1]

    sheet['B47'] = competitors[19][0]
    sheet['B48'] = competitors[19][1]

    sheet['B55'] = competitors[20][0]
    sheet['B56'] = competitors[20][1]

    sheet['B75'] = competitors[21][0]
    sheet['B76'] = competitors[21][1]

    sheet['B107'] = competitors[22][0]
    sheet['B108'] = competitors[22][1]

    sheet['B23'] = competitors[23][0]
    sheet['B24'] = competitors[23][1]

    sheet['B31'] = competitors[24][0]
    sheet['B32'] = competitors[24][1]

    sheet['B99'] = competitors[25][0]
    sheet['B100'] = competitors[25][1]

    sheet['B67'] = competitors[26][0]
    sheet['B68'] = competitors[26][1]

    sheet['B63'] = competitors[27][0]
    sheet['B64'] = competitors[27][1]

    sheet['B39'] = competitors[28][0]
    sheet['B40'] = competitors[28][1]

    sheet['B91'] = competitors[29][0]
    sheet['B92'] = competitors[29][1]

    sheet['B123'] = competitors[30][0]
    sheet['B124'] = competitors[30][1]

    sheet['B7'] = competitors[31][0]
    sheet['B8'] = competitors[31][1]

    ## entering the bye competitors
    sheet['E5'] = competitors_with_byes[0][0]
    sheet['E6'] = competitors_with_byes[0][1]


    sheet['E125'] = competitors_with_byes[0][0]
    sheet['E126'] = competitors_with_byes[0][1]
    
    sheet['E93'] = competitors_with_byes[0][0]
    sheet['E94'] = competitors_with_byes[0][1]

    sheet['E37'] = competitors_with_byes[0][0]
    sheet['E38'] = competitors_with_byes[0][1]

    sheet['E61'] = competitors_with_byes[0][0]
    sheet['E62'] = competitors_with_byes[0][1]

    sheet['E69'] = competitors_with_byes[0][0]
    sheet['E70'] = competitors_with_byes[0][1]

    sheet['E101'] = competitors_with_byes[0][0]
    sheet['E102'] = competitors_with_byes[0][1]

    sheet['E29'] = competitors_with_byes[0][0]
    sheet['E30'] = competitors_with_byes[0][1]

    sheet['E21'] = competitors_with_byes[0][0]
    sheet['E22'] = competitors_with_byes[0][1]

    sheet['E109'] = competitors_with_byes[0][0]
    sheet['E110'] = competitors_with_byes[0][1]

    sheet['E77'] = competitors_with_byes[0][0]
    sheet['E78'] = competitors_with_byes[0][1]

    sheet['E53'] = competitors_with_byes[0][0]
    sheet['E54'] = competitors_with_byes[0][1]

    sheet['E45'] = competitors_with_byes[0][0]
    sheet['E46'] = competitors_with_byes[0][1]

    sheet['E85'] = competitors_with_byes[0][0]
    sheet['E86'] = competitors_with_byes[0][1]

    sheet['E117'] = competitors_with_byes[0][0]
    sheet['E118'] = competitors_with_byes[0][1]
    

    save_location = ask_save_location()
    if save_location is not None:
        workbook.save(save_location)
    else:
        workbook.close()
def process_dataframe(df:pd.DataFrame):
    concatenated_fields = []
    
    for index, row in df.iterrows():
        concatenated_fields.append([row['first_name'] + " " + row['last_name'], row['team']])
    
    total_records = len(df)
    power_of_2 = math.ceil(math.log2(total_records))
    tree_size = 2 ** power_of_2

    num_missing_records = 0
    if total_records < 2 ** power_of_2:
            num_missing_records = 2 ** power_of_2 - total_records
            for i in range(num_missing_records):
                concatenated_fields.append(['Bye', ''])

    bye_list = [["", ""]]*len(concatenated_fields)
    for i in range(num_missing_records):
        bye_list[i] = concatenated_fields[i]
    
    return concatenated_fields, tree_size, bye_list
def generate_xlsx_files():
    randomize_teams_sql = """
    select
    x.*
    from temp_table x
    join 
	(SELECT DISTINCT team, RAND() rnd FROM temp_table GROUP BY team) y ON y.team = x.team
    order by y.rnd, RAND();
    """

    grid_df = get_dataframe(get_connection("pma_tournaments"), randomize_teams_sql)
    id_list = grid_df["id"].tolist()
    event = backend.current_event_filter

    competitors, n, competitors_with_byes = process_dataframe(grid_df)

    if n == 2:
        two_tree(competitors, competitors_with_byes)
    elif n == 4:
        four_tree(competitors, competitors_with_byes)
    elif n == 8:
        eight_tree(competitors, competitors_with_byes)
    elif n == 16:
        sixteen_tree(competitors, competitors_with_byes)
    elif n == 32:
        thirtytwo_tree(competitors, competitors_with_byes)
    else:
        print("There are too many competitors selected, a template of that size does not exist")

    backend.mark_event_complete(event, id_list)